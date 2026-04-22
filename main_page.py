# main_page页面
import tkinter as tk
from tkinter import messagebox, simpledialog
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format

from settings_dialog import SettingsDialog


class MainPage:
    def __init__(self, root, excel_data, file_path, sheet_name, settings):
        self.root = root
        self.excel_data = excel_data
        self.file_path = file_path
        self.sheet_name = sheet_name

        self.current_row = 0
        self.total_patients = len(self.excel_data)
        self.modified_cells = set()

        self.settings = settings.copy()
        self.apply_settings(self.settings)

        self.workbook = load_workbook(self.file_path)
        self.worksheet = self.workbook[self.sheet_name]

        self.colors = {
            "bg": "#f7f8fa",
            "panel": "#f7f8fa",
            "text": "#1f2937",
            "subtext": "#6b7280",
            "neutral_btn": "#eceff3",
            "neutral_btn_active": "#dde3ea",
            "success": "#16a34a",
            "success_active": "#15803d",
            "empty": "#9ca3af",
            "empty_active": "#6b7280"
        }

        self.root.configure(bg=self.colors["bg"])
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.create_patient_workbook()

    def apply_settings(self, settings):
        self.buttons_per_row = max(1, int(settings.get("buttons_per_row", 5)))
        self.visible_rows = max(1, int(settings.get("visible_rows", 6)))
        self.button_text_max = max(1, int(settings.get("button_text_max", 8)))

    def create_patient_workbook(self):
        if self.total_patients == 0:
            messagebox.showerror("错误", "Excel中没有可用数据")
            self.root.destroy()
            return

        self.patient_id = self.excel_data.loc[self.current_row, "住院号"]

        self.workbook_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.workbook_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=14)

        self.workbook_frame.grid_rowconfigure(1, weight=1)
        self.workbook_frame.grid_columnconfigure(0, weight=1)

        self.header_frame = tk.Frame(self.workbook_frame, bg=self.colors["bg"])
        self.header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        self.patient_id_label = tk.Label(
            self.header_frame,
            text="病人ID: {0}".format(self.get_patient_id_text()),
            font=("微软雅黑", 15, "bold"),
            bg=self.colors["bg"],
            fg=self.colors["text"],
            cursor="hand2"
        )
        self.patient_id_label.pack(pady=(0, 6))
        self.patient_id_label.bind("<Button-1>", self.copy_patient_id)

        self.current_patient_label = tk.Label(
            self.header_frame,
            text="当前病人: {0}/{1}".format(self.current_row + 1, self.total_patients),
            font=("微软雅黑", 11),
            bg=self.colors["bg"],
            fg=self.colors["subtext"]
        )
        self.current_patient_label.pack()

        self.buttons_outer_frame = tk.Frame(self.workbook_frame, bg=self.colors["bg"])
        self.buttons_outer_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 8))
        self.buttons_outer_frame.grid_rowconfigure(1, weight=1)
        self.buttons_outer_frame.grid_columnconfigure(0, weight=1)

        self.buttons_title = tk.Label(
            self.buttons_outer_frame,
            text="字段录入区",
            font=("微软雅黑", 11, "bold"),
            bg=self.colors["bg"],
            fg=self.colors["text"],
            anchor="w"
        )
        self.buttons_title.grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.buttons_area_frame = tk.Frame(self.buttons_outer_frame, bg=self.colors["bg"])
        self.buttons_area_frame.grid(row=1, column=0, sticky="nsew")
        self.buttons_area_frame.grid_rowconfigure(0, weight=1)
        self.buttons_area_frame.grid_columnconfigure(0, weight=1)

        self.buttons_canvas = tk.Canvas(
            self.buttons_area_frame,
            bg=self.colors["bg"],
            highlightthickness=0,
            bd=0
        )
        self.buttons_scrollbar = tk.Scrollbar(
            self.buttons_area_frame,
            orient=tk.VERTICAL,
            command=self.buttons_canvas.yview
        )
        self.buttons_canvas.configure(yscrollcommand=self.buttons_scrollbar.set)

        self.buttons_canvas.grid(row=0, column=0, sticky="nsew")
        self.buttons_scrollbar.grid(row=0, column=1, sticky="ns")

        self.buttons_frame = tk.Frame(self.buttons_canvas, bg=self.colors["bg"])
        self.buttons_window = self.buttons_canvas.create_window(
            (0, 0),
            window=self.buttons_frame,
            anchor="nw"
        )

        self.buttons_frame.bind("<Configure>", self.on_buttons_frame_configure)
        self.buttons_canvas.bind("<Configure>", self.on_canvas_configure)
        self.buttons_canvas.bind("<Enter>", self.bind_mousewheel)
        self.buttons_canvas.bind("<Leave>", self.unbind_mousewheel)

        self.navigation_frame = tk.Frame(self.workbook_frame, bg=self.colors["bg"])
        self.navigation_frame.grid(row=2, column=0, sticky="ew", pady=(4, 8))

        self.first_button = self.create_action_button(
            self.navigation_frame, "首页", self.first_patient
        )
        self.first_button.grid(row=0, column=0, padx=4, pady=2)

        self.previous_button = self.create_action_button(
            self.navigation_frame, "上一个", self.previous_patient
        )
        self.previous_button.grid(row=0, column=1, padx=4, pady=2)

        self.jump_button = self.create_action_button(
            self.navigation_frame, "跳转", self.jump_to_patient
        )
        self.jump_button.grid(row=0, column=2, padx=4, pady=2)

        self.next_button = self.create_action_button(
            self.navigation_frame, "下一个", self.next_patient
        )
        self.next_button.grid(row=0, column=3, padx=4, pady=2)

        self.last_button = self.create_action_button(
            self.navigation_frame, "末页", self.last_patient
        )
        self.last_button.grid(row=0, column=4, padx=4, pady=2)

        self.settings_button = self.create_action_button(
            self.navigation_frame, "设置", self.open_settings_window
        )
        self.settings_button.grid(row=0, column=5, padx=4, pady=2)

        self.bottom_frame = tk.Frame(self.workbook_frame, bg=self.colors["bg"])
        self.bottom_frame.grid(row=3, column=0, sticky="ew", pady=(0, 2))

        self.finish_button = self.create_action_button(
            self.bottom_frame, "结束", self.save_and_exit, width=14
        )
        self.finish_button.pack()

        self.button_widgets = {}
        self.render_field_buttons()
        self.update_patient_workbook()
        self.adjust_window_geometry()

    def create_action_button(self, parent, text, command, width=9):
        return tk.Button(
            parent,
            text=text,
            command=command,
            width=width,
            font=("微软雅黑", 10),
            bg=self.colors["neutral_btn"],
            fg=self.colors["text"],
            activebackground=self.colors["neutral_btn_active"],
            activeforeground=self.colors["text"],
            relief="flat",
            bd=0,
            padx=6,
            pady=6,
            cursor="hand2"
        )

    def create_field_button(self, parent, text, command, width):
        return tk.Button(
            parent,
            text=text,
            command=command,
            width=width,
            font=("微软雅黑", 10),
            fg="white",
            relief="flat",
            bd=0,
            padx=6,
            pady=7,
            cursor="hand2"
        )

    def get_button_area_height(self):
        height = self.visible_rows * 46 + 12
        return max(120, min(height, 700))

    def render_field_buttons(self):
        for widget in self.buttons_frame.winfo_children():
            widget.destroy()

        self.button_widgets = {}

        row_index = 0
        col_index = 0
        button_width = max(10, min(self.button_text_max + 2, 18))

        for col_name in self.excel_data.columns:
            if col_name == "住院号":
                continue

            display_text = self.truncate_text(col_name, self.button_text_max)

            button = self.create_field_button(
                self.buttons_frame,
                text=display_text,
                width=button_width,
                command=lambda c=col_name: self.open_input_window(c)
            )
            button.grid(row=row_index, column=col_index, padx=6, pady=6)
            self.button_widgets[col_name] = button
            self.update_button_color(button, col_name)

            col_index += 1
            if col_index >= self.buttons_per_row:
                col_index = 0
                row_index += 1

        self.buttons_frame.update_idletasks()
        content_width = self.buttons_frame.winfo_reqwidth()

        self.buttons_canvas.itemconfigure(self.buttons_window, width=content_width)
        self.refresh_scroll_region()

    def adjust_window_geometry(self):
        self.root.update_idletasks()

        buttons_area_width = self.buttons_frame.winfo_reqwidth() + 28
        header_width = max(
            self.patient_id_label.winfo_reqwidth(),
            self.current_patient_label.winfo_reqwidth()
        )
        nav_width = self.navigation_frame.winfo_reqwidth()
        finish_width = self.finish_button.winfo_reqwidth() + 20

        target_width = max(buttons_area_width, header_width, nav_width, finish_width)
        target_width = target_width + 56
        target_width = max(620, min(target_width, 1500))

        target_height = (
            self.header_frame.winfo_reqheight()
            + self.navigation_frame.winfo_reqheight()
            + self.bottom_frame.winfo_reqheight()
            + self.get_button_area_height()
            + 90
        )
        target_height = max(420, min(target_height, 980))

        min_height = (
            self.header_frame.winfo_reqheight()
            + self.navigation_frame.winfo_reqheight()
            + self.bottom_frame.winfo_reqheight()
            + 140
        )
        min_height = max(320, min_height)

        self.root.geometry("{0}x{1}".format(target_width, target_height))
        self.root.minsize(520, min_height)

    def refresh_scroll_region(self):
        self.buttons_canvas.update_idletasks()
        self.buttons_canvas.configure(scrollregion=self.buttons_canvas.bbox("all"))

    def on_buttons_frame_configure(self, event):
        self.refresh_scroll_region()

    def on_canvas_configure(self, event):
        self.refresh_scroll_region()

    def bind_mousewheel(self, event):
        self.buttons_canvas.bind_all("<MouseWheel>", self.on_mousewheel)
        self.buttons_canvas.bind_all("<Button-4>", self.on_mousewheel_linux_up)
        self.buttons_canvas.bind_all("<Button-5>", self.on_mousewheel_linux_down)

    def unbind_mousewheel(self, event):
        self.buttons_canvas.unbind_all("<MouseWheel>")
        self.buttons_canvas.unbind_all("<Button-4>")
        self.buttons_canvas.unbind_all("<Button-5>")

    def on_mousewheel(self, event):
        if event.delta != 0:
            self.buttons_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_mousewheel_linux_up(self, event):
        self.buttons_canvas.yview_scroll(-1, "units")

    def on_mousewheel_linux_down(self, event):
        self.buttons_canvas.yview_scroll(1, "units")

    def truncate_text(self, text, max_length):
        text = str(text)
        if len(text) <= max_length:
            return text
        return text[:max_length]

    def get_patient_id_text(self):
        value = self.patient_id

        if pd.isna(value):
            return ""

        if isinstance(value, (int, np.integer)):
            return str(int(value))

        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)

        return str(value)

    def has_value(self, value):
        if pd.isna(value):
            return False

        if isinstance(value, str) and value.strip() == "":
            return False

        return True

    def copy_patient_id(self, event=None):
        patient_id_text = self.get_patient_id_text()
        self.root.clipboard_clear()
        self.root.clipboard_append(patient_id_text)
        self.root.update()

    def get_excel_cell(self, row_index, column):
        excel_row = row_index + 2
        excel_col = self.excel_data.columns.get_loc(column) + 1
        return self.worksheet.cell(row=excel_row, column=excel_col)

    def try_parse_date(self, text):
        text = str(text).strip()
        if text == "":
            return None

        date_formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%Y.%m.%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M",
            "%Y.%m.%d %H:%M"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue

        return None

    def try_parse_number(self, text):
        text = str(text).strip()
        if text == "":
            return None

        try:
            if "." in text:
                return float(text)
            return int(text)
        except ValueError:
            return None

    def open_input_window(self, column):
        self.root.attributes("-topmost", False)

        current_value = self.excel_data.at[self.current_row, column]
        if pd.isna(current_value):
            current_value = ""
        else:
            current_value = str(current_value)

        cell = self.get_excel_cell(self.current_row, column)
        cell_format = cell.number_format

        input_value = simpledialog.askstring(
            "输入 {0}".format(column),
            "请输入 {0} 的值:\n原单元格格式: {1}".format(column, cell_format),
            initialvalue=current_value,
            parent=self.root
        )

        self.root.attributes("-topmost", True)

        if input_value is not None:
            self.save_input(column, input_value)

    def save_input(self, column, value):
        if value == "":
            self.excel_data.at[self.current_row, column] = np.nan
        else:
            self.excel_data.at[self.current_row, column] = str(value)

        self.modified_cells.add((self.current_row, column))
        self.update_button_color(self.button_widgets[column], column)

    def first_patient(self):
        if self.current_row != 0:
            self.save_pending_changes()
            self.current_row = 0
            self.update_patient_workbook()

    def last_patient(self):
        last_index = self.total_patients - 1
        if self.current_row != last_index:
            self.save_pending_changes()
            self.current_row = last_index
            self.update_patient_workbook()

    def next_patient(self):
        self.save_pending_changes()
        self.current_row += 1

        if self.current_row < self.total_patients:
            self.update_patient_workbook()
        else:
            self.current_row = self.total_patients - 1
            messagebox.showinfo("完成", "所有患者数据已收集完成！")

    def previous_patient(self):
        if self.current_row > 0:
            self.save_pending_changes()
            self.current_row -= 1
            self.update_patient_workbook()

    def jump_to_patient(self):
        target = simpledialog.askinteger(
            "跳转",
            "请输入要跳转的病人序号（1 - {0}）:".format(self.total_patients),
            parent=self.root,
            minvalue=1,
            maxvalue=self.total_patients
        )

        if target is None:
            return

        target_index = target - 1

        if target_index != self.current_row:
            self.save_pending_changes()
            self.current_row = target_index
            self.update_patient_workbook()

    def open_settings_window(self):
        dialog = SettingsDialog(
            self.root,
            {
                "buttons_per_row": self.buttons_per_row,
                "visible_rows": self.visible_rows,
                "button_text_max": self.button_text_max
            }
        )
        result = dialog.show()

        if result is None:
            return

        self.settings = result.copy()
        self.apply_settings(self.settings)
        self.render_field_buttons()
        self.update_patient_workbook()
        self.adjust_window_geometry()

    def update_patient_workbook(self):
        self.patient_id = self.excel_data.loc[self.current_row, "住院号"]
        self.root.title("佰态 · 病例数据采集工作台 - {0}".format(self.get_patient_id_text()))

        self.patient_id_label.config(
            text="病人ID: {0}".format(self.get_patient_id_text())
        )
        self.current_patient_label.config(
            text="当前病人: {0}/{1}".format(self.current_row + 1, self.total_patients)
        )

        for col_name, button in self.button_widgets.items():
            self.update_button_color(button, col_name)

    def save_pending_changes(self):
        if not self.modified_cells:
            return

        sorted_cells = sorted(
            self.modified_cells,
            key=lambda item: (
                item[0],
                self.excel_data.columns.get_loc(item[1])
            )
        )

        for row_index, column in sorted_cells:
            self.write_single_cell(row_index, column)

        self.workbook.save(self.file_path)
        self.modified_cells.clear()

    def write_single_cell(self, row_index, column):
        cell = self.get_excel_cell(row_index, column)
        value = self.excel_data.at[row_index, column]

        if pd.isna(value) or value == "":
            cell.value = None
            return

        original_format = cell.number_format
        original_value = cell.value

        # 如果原单元格是日期格式，优先按日期写入
        if is_date_format(original_format):
            parsed_date = self.try_parse_date(value)
            if parsed_date is not None:
                cell.value = parsed_date
                return

        # 如果原单元格本来是数字，优先按数字写入
        if isinstance(original_value, (int, float)):
            parsed_number = self.try_parse_number(value)
            if parsed_number is not None:
                cell.value = parsed_number
                return

        # 其他情况按文本写入
        cell.value = str(value)

    def update_button_color(self, button, column):
        current_value = self.excel_data.at[self.current_row, column]

        if self.has_value(current_value):
            button.config(
                bg=self.colors["success"],
                activebackground=self.colors["success_active"],
                activeforeground="white"
            )
        else:
            button.config(
                bg=self.colors["empty"],
                activebackground=self.colors["empty_active"],
                activeforeground="white"
            )

    def save_and_exit(self):
        self.save_pending_changes()
        messagebox.showinfo("保存成功", "已保存你的劳动成果，休息吧，小牛马！")
        self.root.destroy()

    def on_closing(self):
        self.save_pending_changes()
        messagebox.showinfo("保存成功", "已保存你的劳动成果，休息吧，小牛马！")
        self.root.destroy()